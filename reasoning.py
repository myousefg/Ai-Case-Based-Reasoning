# Import library - pandas untuk membaca file restoran.xlsx
import pandas as pd

# Fuzzification - menggunakan fungsi triangular untuk menentukan nilai keanggotaan.
def triangular(x, a, b, c):
    if a == b and b == c:
        return 1.0 if x == a else 0.0
    if x == b:
        return 1.0
    if x <= a or x >= c:
        return 0.0
    if a == b:
        return (c - x) / (c - b)
    if b == c:
        return (x - a) / (b - a)
    if x < b:
        return (x - a) / (b - a)
    else:
        return (c - x) / (c - b)

# Definisi batas Kualitas Servis
def servis_membership(x):
    return {
        'Buruk': triangular(x, 0, 20, 40),
        'Sedang': triangular(x, 30, 50, 70),
        'Bagus': triangular(x, 60, 80, 100)
    }

# Definisi batas Harga
def harga_membership(x):
    return {
        'Murah': triangular(x, 25000, 25000, 35000),
        'Sedang': triangular(x, 30000, 40000, 50000),
        'Mahal': triangular(x, 45000, 55000, 55000)
    }

# Output fuzzy sets for eligibility
output_terms = {
    'Rendah': (0, 0, 50),
    'Sedang': (0, 50, 100),
    'Tinggi': (50, 100, 100)
}

# Fuzzyfication - Definisi aturan fuzzy
rules = [
    ('Bagus', 'Murah', 'Tinggi'),
    ('Bagus', 'Sedang', 'Tinggi'),
    ('Bagus', 'Mahal', 'Sedang'),
    ('Sedang', 'Murah', 'Tinggi'),
    ('Sedang', 'Sedang', 'Sedang'),
    ('Sedang', 'Mahal', 'Rendah'),
    ('Buruk', 'Murah', 'Sedang'),
    ('Buruk', 'Sedang', 'Rendah'),
    ('Buruk', 'Mahal', 'Rendah')
]

def compute_eligibility_score(servis_value, harga_value):
    m_serv = servis_membership(servis_value)
    m_harga = harga_membership(harga_value)

    output_strength = {'Rendah': 0.0, 'Sedang': 0.0, 'Tinggi': 0.0}
    
    for serv_term, hrg_term, out_term in rules:
        strength = min(m_serv[serv_term], m_harga[hrg_term])
        if strength > output_strength[out_term]:
            output_strength[out_term] = strength

    # Defuzzification - menggunakan method centroid 
    x_values = list(range(0, 101))
    numerator = 0.0
    denominator = 0.0
    
    for x in x_values:
        memberships = []
        for term, (a, b, c) in output_terms.items():
            base = triangular(x, a, b, c)
            clipped = min(output_strength[term], base)
            memberships.append(clipped)
        agg = max(memberships)
        numerator += x * agg
        denominator += agg

    if denominator == 0:
        return 0.0
    return numerator / denominator

# Main program
if __name__ == "__main__":
    import openpyxl
    from openpyxl.styles import Border, Side, Alignment

    df = pd.read_excel('restoran.xlsx')

    df['Kelayakan Skor'] = [compute_eligibility_score(r['Pelayanan'], r['harga']) for _, r in df.iterrows()]
    df_sorted = df.sort_values(by='Kelayakan Skor', ascending=False)

    top5 = df_sorted.head(5)[['id Pelanggan', 'Pelayanan', 'harga', 'Kelayakan Skor']]
    top5.columns = ['Id', 'Kualitas Servis', 'Harga', 'Kelayakan Skor']
    top5.reset_index(drop=True, inplace=True)
    top5.insert(0, 'No.', top5.index + 1)

    top5.to_excel('peringkat.xlsx', index=False)

    wb = openpyxl.load_workbook('peringkat.xlsx')
    ws = wb.active
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border, cell.alignment = border, align_center

    wb.save('peringkat.xlsx')
    print("Top 5 restaurants saved to peringkat.xlsx with centered text and borders.")